import os, sys, time, pathlib, re, requests, datetime
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup

import pdfplumber
import pandas as pd

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from zoneinfo import ZoneInfo

# ---------------------- CONFIG ----------------------

NALCO_URL = "https://nalcoindia.com/domestic/current-price/"
PDF_DIR = pathlib.Path("pdfs")
DATA_DIR = pathlib.Path("data")
LOG_FILE = DATA_DIR / "latest_nalco_pdf.txt"
EXCEL_FILE = DATA_DIR / "nalco_prices.xlsx"
RUNLOG_FILE = DATA_DIR / "nalco_run_log.xlsx"

UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36")

# prefer "Ingot-DD-MM-YYYY.pdf" and exclude spec docs
DATEY_PDF_RE = re.compile(r"Ingot-(\d{2})-(\d{2})-(\d{4})\.pdf$", re.IGNORECASE)

# Final Excel column order
EXCEL_COLS = ["Sl.no.", "Description", "Product Code", "Basic Price", "Circular Date", "Circular Link"]

# ---------------------- SCRAPER UTILS ----------------------

def ensure_dirs():
    for p in (PDF_DIR, DATA_DIR):
        if p.exists() and p.is_file():
            p.unlink()
        p.mkdir(parents=True, exist_ok=True)

def get_html(url):
    r = requests.get(url, headers={"User-Agent": UA}, timeout=60)
    r.raise_for_status()
    return r.text

def norm(s: str) -> str:
    return (s or "").strip().lower()

def find_ingots_pdf_url(html):
    soup = BeautifulSoup(html, "html.parser")

    # STRICT: <a ...><img ...><p>Ingots</p></a>
    for pnode in soup.find_all("p"):
        if norm(pnode.get_text()) == "ingots":
            a = pnode.find_parent("a", href=True)
            if a:
                href = a["href"].strip()
                if href.lower().endswith(".pdf") and "spec" not in href.lower():
                    return urljoin(NALCO_URL, href)

    # Prefer filenames like Ingot-DD-MM-YYYY.pdf (exclude spec)
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().endswith(".pdf") and "spec" not in href.lower():
            if DATEY_PDF_RE.search(href):
                return urljoin(NALCO_URL, href)

    # Fallback: any PDF link whose anchor text mentions "ingot" (not spec)
    for a in soup.find_all("a", href=True):
        txt = norm(a.get_text())
        href = a["href"].strip()
        if href.lower().endswith(".pdf") and "ingot" in txt and "spec" not in href.lower():
            return urljoin(NALCO_URL, href)

    return None

def read_last_url():
    return LOG_FILE.read_text(encoding="utf-8").strip() if LOG_FILE.exists() else ""

def write_last_url(url):
    LOG_FILE.write_text(url or "", encoding="utf-8")

def download_pdf(url):
    headers = {
        "User-Agent": UA,
        "Referer": NALCO_URL,
        "Accept": "application/pdf,*/*;q=0.9",
        "Accept-Language": "en-US,en;q=0.9",
    }
    with requests.get(url, headers=headers, timeout=60, stream=True, allow_redirects=True) as r:
        r.raise_for_status()
        ctype = r.headers.get("Content-Type", "").lower()
        if "application/pdf" not in ctype:
            raise RuntimeError(f"Expected PDF but got Content-Type={ctype!r} from {url}")
        filename = None
        cd = r.headers.get("Content-Disposition", "")
        if "filename=" in cd:
            filename = cd.split("filename=", 1)[1].strip('"; ')
        if not filename:
            filename = os.path.basename(urlparse(r.url).path) or f"nalco_{int(time.time())}.pdf"
        dest = PDF_DIR / filename
        with open(dest, "wb") as f:
            for chunk in r.iter_content(chunk_size=65536):
                if chunk:
                    f.write(chunk)
        return dest

# ---------------------- PDF PARSING ----------------------

def parse_circular_date_from_filename(pdf_path: pathlib.Path) -> str:
    """Extract DD-MM-YYYY from 'Ingot-DD-MM-YYYY.pdf', else use today."""
    m = DATEY_PDF_RE.search(pdf_path.name)
    if m:
        dd, mm, yyyy = m.groups()
        try:
            d = datetime.date(int(yyyy), int(mm), int(dd))
            return d.strftime("%d-%m-%Y")
        except ValueError:
            pass
    return datetime.date.today().strftime("%d-%m-%Y")

def extract_row_ie07(pdf_path: pathlib.Path):
    """
    Find the row that contains 'IE07'.
    Returns dict: { 'Description', 'Product Code', 'Basic Price' }
    """
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # Try structured tables
            try:
                tables = page.extract_tables()
            except Exception:
                tables = []
            for tbl in tables or []:
                for row in tbl:
                    cells = [(c or "").strip() for c in row]
                    if any(re.fullmatch(r"IE07", x, flags=re.IGNORECASE) for x in cells):
                        desc = ""
                        code = "IE07"
                        price = ""
                        idx_code = None
                        for i, c in enumerate(cells):
                            if re.fullmatch(r"IE07", c, flags=re.IGNORECASE):
                                idx_code = i
                                break
                        if idx_code is not None:
                            # description: look left for non-numeric text
                            for j in range(idx_code - 1, -1, -1):
                                if cells[j] and not re.fullmatch(r"\d+(\.\d+)?", cells[j]):
                                    desc = cells[j]
                                    break
                            # price: first numeric-looking to right
                            for j in range(idx_code + 1, len(cells)):
                                if re.search(r"\d", cells[j]):
                                    price = cells[j].replace(",", "")
                                    break
                        if code and price:
                            return {
                                "Description": (desc or "ALUMINIUM INGOT").upper(),
                                "Product Code": code,
                                "Basic Price": price
                            }

            # Fallback: line text scan
            words = page.extract_words(use_text_flow=True, keep_blank_chars=False)
            lines = {}
            for w in words:
                y = round(w["top"], 1)
                lines.setdefault(y, []).append(w)
            for y, wlist in lines.items():
                text_line = " ".join([w["text"] for w in sorted(wlist, key=lambda x: x["x0"])])
                if re.search(r"\bIE07\b", text_line, flags=re.IGNORECASE):
                    m_price = re.search(r"(\d{5,7}(?:\.\d+)?)\s*$", text_line)
                    price = (m_price.group(1) if m_price else "").replace(",", "")
                    m_desc = re.search(r"([A-Z ]*INGOT[A-Z ]*)\b", text_line, flags=re.IGNORECASE)
                    desc = (m_desc.group(1) if m_desc else "ALUMINIUM INGOT").strip().upper()
                    if price:
                        return {
                            "Description": desc,
                            "Product Code": "IE07",
                            "Basic Price": price
                        }
    raise RuntimeError("Could not find a row with Product Code IE07 in the PDF.")

def to_thousands(value_str: str) -> float:
    """Convert raw price (e.g., '268250') to thousands (e.g., 268.250)."""
    value_str = value_str.replace(",", "").strip()
    if not value_str:
        return None
    try:
        v = float(value_str)
        return round(v / 1000.0, 3)
    except ValueError:
        return None

# ---------------------- EXCEL HELPERS ----------------------

def sort_and_format_df(df: pd.DataFrame) -> pd.DataFrame:
    """Sort by Circular Date (desc) for display. Keep Basic Price numeric (3 decimals)."""
    dtd = pd.to_datetime(df["Circular Date"], dayfirst=True, errors="coerce")
    df = df.assign(_date=dtd)
    df = df.sort_values(by=["_date", "Sl.no."], ascending=[False, True], kind="stable").drop(columns=["_date"])
    df["Basic Price"] = pd.to_numeric(df["Basic Price"], errors="coerce").round(3)
    df["Circular Date"] = pd.to_datetime(df["Circular Date"], dayfirst=True, errors="coerce").dt.strftime("%d-%m-%Y")
    return df

def save_excel_formatted(df: pd.DataFrame, path: pathlib.Path):
    """
    Save df to Excel, auto-fit column widths based on content,
    center-align all cells, and make Circular Link clickable.
    """
    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    center = Alignment(horizontal="center", vertical="center")

    # Auto width
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name].astype(str).values:
            if val is None:
                continue
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 80))

    header_row = 1
    nrows = ws.max_row
    ncols = ws.max_column
    price_col_idx = EXCEL_COLS.index("Basic Price") + 1
    link_col_idx = EXCEL_COLS.index("Circular Link") + 1

    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center
            if r > header_row and c == price_col_idx:
                cell.number_format = "0.000"
            if r > header_row and c == link_col_idx:
                val = cell.value
                if isinstance(val, str) and val.startswith("http"):
                    cell.hyperlink = val

    ws.freeze_panes = "A2"
    wb.save(path)

def append_to_excel(excel_path: pathlib.Path, row: dict) -> int:
    """
    Append a row, assign next Sl.no. based on existing max (robust even after sorting),
    then sort by Circular Date (desc) before saving with formatting.
    Returns total rows after append.
    """
    if excel_path.exists():
        df = pd.read_excel(excel_path, dtype={"Sl.no.": "Int64"})
        for c in EXCEL_COLS:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[EXCEL_COLS]
        next_slno = int(df["Sl.no."].max()) + 1 if df["Sl.no."].notna().any() else 1
    else:
        df = pd.DataFrame(columns=EXCEL_COLS)
        next_slno = 1

    new_row = {
        "Sl.no.": next_slno,
        "Description": row["Description"],
        "Product Code": row["Product Code"],
        "Basic Price": row["Basic Price"],    # already divided by 1000 upstream
        "Circular Date": row["Circular Date"],
        "Circular Link": row["Circular Link"],
    }

    df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    df = sort_and_format_df(df)
    save_excel_formatted(df, excel_path)
    return df.shape[0]

# ---------------------- RUN LOG HELPERS ----------------------

def append_runlog(log_path: pathlib.Path, info: dict):
    """
    Append (or create) a run log Excel separate from the main data file.
    Columns: Run UTC, Run IST, Status, Message, Chosen URL, Saved PDF, Rows Appended, Total Rows After
    """
    cols = ["Run UTC", "Run IST", "Status", "Message", "Chosen URL", "Saved PDF", "Rows Appended", "Total Rows After"]
    if log_path.exists():
        df = pd.read_excel(log_path)
        for c in cols:
            if c not in df.columns:
                df[c] = pd.NA
        df = df[cols]
    else:
        df = pd.DataFrame(columns=cols)

    df = pd.concat([df, pd.DataFrame([{
        "Run UTC": info.get("Run UTC"),
        "Run IST": info.get("Run IST"),
        "Status": info.get("Status"),
        "Message": info.get("Message"),
        "Chosen URL": info.get("Chosen URL"),
        "Saved PDF": info.get("Saved PDF"),
        "Rows Appended": info.get("Rows Appended"),
        "Total Rows After": info.get("Total Rows After"),
    }])], ignore_index=True)

    # Save & simple formatting (autofit + center)
    df.to_excel(log_path, index=False)
    wb = load_workbook(log_path)
    ws = wb.active
    center = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for val in df[col_name].astype(str).values:
            max_len = max(max_len, len(val))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(max_len + 2, 100))
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).alignment = center
    ws.freeze_panes = "A2"
    wb.save(log_path)

def now_times():
    now_utc = datetime.datetime.now(datetime.timezone.utc)
    now_ist = now_utc.astimezone(ZoneInfo("Asia/Kolkata"))
    return now_utc.strftime("%Y-%m-%d %H:%M:%S UTC"), now_ist.strftime("%Y-%m-%d %H:%M:%S IST")

# ---------------------- MAIN FLOW ----------------------

def main():
    ensure_dirs()
    run_utc, run_ist = now_times()

    html = get_html(NALCO_URL)
    pdf_url = find_ingots_pdf_url(html)
    if not pdf_url:
        msg = "No Ingots PDF link found on the page."
        print(msg, file=sys.stderr)
        append_runlog(RUNLOG_FILE, {
            "Run UTC": run_utc, "Run IST": run_ist, "Status": "SKIPPED",
            "Message": msg, "Chosen URL": "", "Saved PDF": "", "Rows Appended": 0, "Total Rows After": ""
        })
        sys.exit(1)

    print(f"Chosen PDF URL: {pdf_url}")

    last = read_last_url()
    if pdf_url == last:
        msg = "No change in PDF. Skipping download & Excel update."
        print(msg)
        append_runlog(RUNLOG_FILE, {
            "Run UTC": run_utc, "Run IST": run_ist, "Status": "SKIPPED",
            "Message": msg, "Chosen URL": pdf_url, "Saved PDF": "", "Rows Appended": 0, "Total Rows After": ""
        })
        return

    # New circular
    pdf_path = download_pdf(pdf_url)
    write_last_url(pdf_url)
    print(f"Saved to: {pdf_path}")

    # Extract IE07 row from PDF
    row = extract_row_ie07(pdf_path)

    # Convert Basic Price to thousands (e.g., 268250 -> 268.250)
    thousands = to_thousands(row["Basic Price"])
    if thousands is None:
        raise RuntimeError(f"Could not parse numeric price from: {row['Basic Price']!r}")
    row["Basic Price"] = thousands

    # Circular date: from filename (fallback to today)
    row["Circular Date"] = parse_circular_date_from_filename(pdf_path)
    # Link: the PDF URL used
    row["Circular Link"] = pdf_url

    total_rows = append_to_excel(EXCEL_FILE, row)
    print(f"Excel updated: {EXCEL_FILE}")

    append_runlog(RUNLOG_FILE, {
        "Run UTC": run_utc, "Run IST": run_ist, "Status": "UPDATED",
        "Message": "New circular processed.", "Chosen URL": pdf_url,
        "Saved PDF": pdf_path.name, "Rows Appended": 1, "Total Rows After": total_rows
    })

if __name__ == "__main__":
    main()
